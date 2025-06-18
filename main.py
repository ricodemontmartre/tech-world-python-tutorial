import openpyxl

inv_file = openpyxl.load_workbook("inventory.xlsx")
print(inv_file["Sheet1"])
product_list = inv_file["Sheet1"]

products_by_supplier = {}
total_value_per_supplier = {}
products_under_10_inv = {}

print(product_list.max_row)

for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_number = product_list.cell(product_row, 1).value
    print(supplier_name)

    # Number of products by supplier
    if supplier_name in products_by_supplier:
        current_num_product = products_by_supplier.get(supplier_name)
        products_by_supplier[supplier_name] = current_num_product + 1
    else:
        print("new supplier added")
        products_by_supplier[supplier_name] = 1

    # Total value of products by suppliers
    if supplier_name in total_value_per_supplier:
        current_total_value = total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[supplier_name] = total_value = current_total_value + price * inventory
    else:
        total_value_per_supplier[supplier_name] = price * inventory


    # Calculation inventory less 10
    if inventory < 10:
        products_under_10_inv[int(product_number)] = int(inventory)


print(products_by_supplier)
print(total_value_per_supplier)
print(products_under_10_inv)